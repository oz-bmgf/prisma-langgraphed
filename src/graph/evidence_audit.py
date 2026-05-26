"""evidence_audit graph — post-hoc diagnostic analysis of a completed analyst run (§13).

Topology (linear, no fan-out, no interrupts):
  START → load_artifacts → run_audit → write_brief
        → write_workbook → write_diagnosis → rollup → END

write_workbook and write_diagnosis are no-ops when their config flags are False.
rollup is a no-op unless program == "all" or a cross-program summary is needed.
"""
from __future__ import annotations

import asyncio
import json
import logging
from collections import Counter, defaultdict
from pathlib import Path
from typing import Optional

from langgraph.graph import END, START, StateGraph

from src.graph.state import EvidenceAuditState

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Audit sub-functions (pure computation — no LLM, no LangGraph imports)
# ---------------------------------------------------------------------------

SEVERITY_WEIGHTS: dict[str, float] = {
    "program_critical": 3.0,
    "pathway_altering": 2.0,
    "efficiency_reducing": 1.0,
    "high": 2.0,
    "medium": 1.0,
    "low": 0.5,
}

CORE_DOC_REQUIREMENTS: list[tuple[str, set[str]]] = [
    ("proposal_or_investment_doc", {"proposal", "investment_document"}),
    ("progress_report",            {"progress_report"}),
    ("budget",                     {"budget"}),
    ("final_or_lifecycle_report",  {"final_report", "lifecycle"}),
]


def _iter_findings(analyst_report: dict) -> list[dict]:
    fs: list[dict] = []
    fs.extend(analyst_report.get("all_findings") or [])
    fs.extend(analyst_report.get("cross_cutting") or [])
    return fs


def _iter_evidence(finding: dict):
    for ev in finding.get("evidence_for") or []:
        yield "for", ev
    for ev in finding.get("evidence_against") or []:
        yield "against", ev


def _cited_file_ids(analyst_report: dict) -> set[str]:
    cited: set[str] = set()
    for f in _iter_findings(analyst_report):
        for _, ev in _iter_evidence(f):
            fid = ev.get("file_id") or ""
            if fid:
                cited.add(fid)
    return cited


def _coverage_analysis(analyst_report: dict, doc_list: list[dict]) -> dict:
    """Ingested docs that were never cited in any finding."""
    cited = _cited_file_ids(analyst_report)

    # Build set of file_ids read (from threads[].documents_read)
    documents_read: set[str] = set()
    for t in analyst_report.get("threads") or []:
        documents_read.update(t.get("documents_read") or [])

    never_cited: list[dict] = []
    read_but_uncited: list[dict] = []
    cited_count = 0

    for d in doc_list:
        fid = d.get("file_id") or ""
        if not fid:
            continue
        if fid in cited:
            cited_count += 1
            continue
        rec = {
            "file_id": fid,
            "filename": d.get("filename", ""),
            "doc_type": d.get("doc_type", ""),
            "inv_id": d.get("inv_id", ""),
            "total_pages": d.get("total_pages", 0),
        }
        if fid in documents_read:
            read_but_uncited.append(rec)
        else:
            never_cited.append(rec)

    by_doc_type = Counter(d["doc_type"] or "unknown" for d in never_cited)
    return {
        "total_docs": len(doc_list),
        "cited_count": cited_count,
        "documents_read_count": len(documents_read),
        "never_cited_count": len(never_cited),
        "read_but_uncited_count": len(read_but_uncited),
        "never_cited_by_doc_type": dict(by_doc_type.most_common()),
        "never_cited": never_cited,
        "read_but_uncited": read_but_uncited,
    }


def _file_influence(analyst_report: dict, doc_index: dict[str, dict], top_n: int = 25) -> list[dict]:
    """Rank files by severity-weighted citation count."""
    counts: dict[str, dict] = defaultdict(lambda: {
        "file_id": "",
        "filename": "",
        "doc_type": "",
        "inv_id": "",
        "citations": 0,
        "weighted_score": 0.0,
        "finding_types": Counter(),
        "finding_ids": set(),
    })

    for f in _iter_findings(analyst_report):
        ftype = f.get("finding_type", "?")
        sev = f.get("severity", "")
        weight = SEVERITY_WEIGHTS.get(sev, 1.0)
        fid_finding = f.get("id", "")
        for _, ev in _iter_evidence(f):
            key = ev.get("file_id") or ""
            if not key:
                continue
            slot = counts[key]
            if not slot["file_id"]:
                slot["file_id"] = key
                slot["filename"] = ev.get("filename", "")
                slot["inv_id"] = ev.get("inv_id", "")
                doc = doc_index.get(key, {})
                slot["doc_type"] = doc.get("doc_type", "")
                if not slot["inv_id"]:
                    slot["inv_id"] = doc.get("inv_id", "")
            slot["citations"] += 1
            slot["weighted_score"] += weight
            slot["finding_types"][ftype] += 1
            slot["finding_ids"].add(fid_finding)

    rows = [
        {
            **{k: v for k, v in slot.items() if k not in ("finding_types", "finding_ids")},
            "finding_types": dict(slot["finding_types"]),
            "finding_ids": sorted(slot["finding_ids"]),
            "distinct_findings": len(slot["finding_ids"]),
        }
        for slot in counts.values()
    ]
    rows.sort(key=lambda r: (r["weighted_score"], r["citations"]), reverse=True)
    return rows[:top_n]


def _source_type_matrix(analyst_report: dict) -> dict:
    """source_type × finding_type cross-tab."""
    matrix: dict[str, Counter] = defaultdict(Counter)
    finding_totals: Counter = Counter()
    source_totals: Counter = Counter()
    for f in _iter_findings(analyst_report):
        ftype = f.get("finding_type", "?")
        for _, ev in _iter_evidence(f):
            stype = (ev.get("source_type") or "?").upper().replace("_", " ").strip()
            matrix[ftype][stype] += 1
            finding_totals[ftype] += 1
            source_totals[stype] += 1
    return {
        "matrix": {k: dict(v) for k, v in matrix.items()},
        "finding_totals": dict(finding_totals),
        "source_totals": dict(source_totals),
    }


def _doc_type_matrix(analyst_report: dict, doc_index: dict[str, dict]) -> dict:
    """doc_type × finding_type cross-tab."""
    matrix: dict[str, Counter] = defaultdict(Counter)
    finding_totals: Counter = Counter()
    doc_totals: Counter = Counter()
    for f in _iter_findings(analyst_report):
        ftype = f.get("finding_type", "?")
        for _, ev in _iter_evidence(f):
            fid = ev.get("file_id") or ""
            doc = doc_index.get(fid, {})
            dtype = doc.get("doc_type") or "unknown"
            matrix[ftype][dtype] += 1
            finding_totals[ftype] += 1
            doc_totals[dtype] += 1
    return {
        "matrix": {k: dict(v) for k, v in matrix.items()},
        "finding_totals": dict(finding_totals),
        "doc_type_totals": dict(doc_totals),
    }


def _weak_findings(analyst_report: dict) -> list[dict]:
    """Findings with low confidence, thin evidence, or against ≥ for."""
    weak_confidences = {"low", "very_low", "moderate", ""}
    out = []
    for f in _iter_findings(analyst_report):
        ev_for = f.get("evidence_for") or []
        ev_against = f.get("evidence_against") or []
        conf = (f.get("confidence") or "").lower()
        is_weak = (
            conf in weak_confidences
            or len(ev_for) <= 1
            or (len(ev_against) > 0 and len(ev_against) >= len(ev_for))
        )
        if not is_weak:
            continue
        out.append({
            "id": f.get("id", ""),
            "statement": (f.get("finding") or f.get("statement") or "")[:400],
            "finding_type": f.get("finding_type", ""),
            "severity": f.get("severity", ""),
            "confidence": conf,
            "evidence_for_count": len(ev_for),
            "evidence_against_count": len(ev_against),
        })
    return out


def _unresolved_gaps(analyst_report: dict) -> list[dict]:
    out = []
    for t in analyst_report.get("threads") or []:
        for g in t.get("gaps") or []:
            entry = {
                "thread_id": t.get("id", ""),
                **(g if isinstance(g, dict) else {"text": str(g)}),
            }
            out.append(entry)
    for r in analyst_report.get("recommended_further_research") or []:
        entry = {"source": "recommended_further_research"}
        if isinstance(r, dict):
            entry.update(r)
        else:
            entry["text"] = str(r)
        out.append(entry)
    return out


def _per_investment_doc_coverage(doc_list: list[dict]) -> dict:
    by_inv: dict[str, Counter] = defaultdict(Counter)
    for d in doc_list:
        inv = d.get("inv_id") or ""
        if not inv:
            continue
        dtype = d.get("doc_type") or "unknown"
        by_inv[inv][dtype] += 1

    n_reqs = len(CORE_DOC_REQUIREMENTS)
    rows = []
    for inv_id, counts in by_inv.items():
        present = {dt for dt, n in counts.items() if n > 0}
        missing = [label for label, satisfiers in CORE_DOC_REQUIREMENTS
                   if not (satisfiers & present)]
        rows.append({
            "inv_id": inv_id,
            "doc_count": sum(counts.values()),
            "doc_types_present": dict(counts),
            "missing_core_types": missing,
            "completeness": (n_reqs - len(missing)) / n_reqs,
        })
    rows.sort(key=lambda r: (len(r["missing_core_types"]), -r["doc_count"]), reverse=True)
    return {"core_doc_types": [label for label, _ in CORE_DOC_REQUIREMENTS], "investments": rows}


def _format_brief(audit: dict) -> str:
    """Render a markdown team brief from the audit dict."""
    program = audit.get("program", "?")
    summary = audit.get("summary", {})
    lines = [
        f"# Evidence Audit Brief — {program}",
        "",
        "## Summary",
        f"- Total findings: {summary.get('total_findings', 0)}",
        f"- Documents available: {summary.get('documents_available', '?')}",
        f"- Evidence quality grade: {summary.get('evidence_quality_grade', '?')}",
        f"- Coverage: {summary.get('coverage_pct', '?')}",
        "",
    ]

    cov = audit.get("coverage", {})
    lines += [
        "## Document Coverage",
        f"- Total ingested docs: {cov.get('total_docs', 0)}",
        f"- Cited in at least one finding: {cov.get('cited_count', 0)}",
        f"- Never cited: {cov.get('never_cited_count', 0)}",
        f"- Read but not cited: {cov.get('read_but_uncited_count', 0)}",
        "",
    ]

    top_files = audit.get("top_files", [])[:10]
    if top_files:
        lines += ["## Top Cited Files (by severity-weighted score)", ""]
        lines.append("| Score | File | Doc Type | Citations |")
        lines.append("|-------|------|----------|-----------|")
        for f in top_files:
            lines.append(
                f"| {f['weighted_score']:.1f} | {f['filename'][:60]} "
                f"| {f['doc_type']} | {f['citations']} |"
            )
        lines.append("")

    weak = audit.get("weak_findings", [])
    if weak:
        lines += [
            f"## Weakly-Evidenced Findings ({len(weak)})",
            "",
        ]
        for w in weak[:15]:
            lines.append(
                f"- **{w['id']}** [{w['severity']}] {w['statement'][:120]}... "
                f"(confidence={w['confidence']}, for={w['evidence_for_count']}, "
                f"against={w['evidence_against_count']})"
            )
        lines.append("")

    gaps = audit.get("unresolved_gaps", [])
    if gaps:
        lines += [f"## Unresolved Gaps ({len(gaps)})", ""]
        for g in gaps[:10]:
            text = g.get("text") or g.get("question") or str(g)[:120]
            lines.append(f"- {text}")
        lines.append("")

    inv_cov = audit.get("investment_doc_coverage", {})
    incomplete = [r for r in inv_cov.get("investments", []) if r["missing_core_types"]]
    if incomplete:
        lines += [f"## Investments Missing Core Doc Types ({len(incomplete)})", ""]
        for r in incomplete[:10]:
            lines.append(f"- {r['inv_id']}: missing {r['missing_core_types']}")
        lines.append("")

    return "\n".join(lines)


def _strip_internal_keys(obj: object) -> object:
    if isinstance(obj, dict):
        return {k: _strip_internal_keys(v) for k, v in obj.items()
                if not (isinstance(k, str) and k.startswith("_"))}
    if isinstance(obj, list):
        return [_strip_internal_keys(v) for v in obj]
    return obj


# ---------------------------------------------------------------------------
# Graph nodes
# ---------------------------------------------------------------------------


async def load_artifacts(state: EvidenceAuditState) -> dict:
    data_root = Path(state["data_root"])
    program = state["program"]
    run_dir_name = state.get("run_dir_name") or "phaseA-data-prep-full"

    experiments_dir = data_root / f"{program}-experiments"
    run_dir = experiments_dir / run_dir_name
    threads_dir = run_dir / "threads"
    ingested_dir = data_root / f"{program}-ingested"

    def _read(path: Path) -> object:
        if not path.exists():
            return None
        try:
            return json.loads(path.read_text(encoding="utf-8"))
        except (json.JSONDecodeError, OSError):
            return None

    # asyncio-APPROVED-1: to_thread wraps blocking JSON file read
    analyst_report = await asyncio.to_thread(_read, threads_dir / "analyst_report.json")
    # asyncio-APPROVED-1: to_thread wraps blocking JSON file read
    doc_list = await asyncio.to_thread(_read, ingested_dir / "doc_list.json")
    # asyncio-APPROVED-1: to_thread wraps blocking JSON file read
    investment_scoring = await asyncio.to_thread(_read, ingested_dir / "investment_scoring.json")

    if analyst_report is None:
        available = sorted(p.name for p in experiments_dir.iterdir() if p.is_dir()) if experiments_dir.is_dir() else []
        return {
            "errors": [
                f"analyst_report.json not found at {threads_dir}. "
                f"Available run dirs: {available}"
            ]
        }

    return {
        "analyst_report": analyst_report,
        "doc_list": doc_list or [],
        "investment_scoring": investment_scoring or {},
        "run_dir": str(run_dir),
    }


async def run_audit(state: EvidenceAuditState) -> dict:
    analyst_report: dict = state.get("analyst_report") or {}
    doc_list: list[dict] = state.get("doc_list") or []
    investment_scoring: dict = state.get("investment_scoring") or {}
    run_dir = state.get("run_dir") or ""
    top_n = state.get("top_n_files") or 25

    doc_index = {d.get("file_id", ""): d for d in doc_list if d.get("file_id")}

    findings = analyst_report.get("all_findings") or []
    cross_cutting = analyst_report.get("cross_cutting") or []

    audit: dict = {
        "program": state["program"],
        "run_dir": run_dir,
        "summary": {
            "evidence_quality_grade": analyst_report.get("evidence_quality_grade"),
            "coverage_pct": analyst_report.get("coverage_pct"),
            "documents_available": analyst_report.get("documents_available"),
            "documents_read": analyst_report.get("documents_read", 0),
            "total_findings": len(findings),
            "cross_cutting_findings": len(cross_cutting),
            "finding_type_counts": dict(Counter(f.get("finding_type", "?") for f in findings)),
            "severity_counts": dict(Counter(f.get("severity", "") for f in findings)),
            "confidence_counts": dict(Counter(f.get("confidence", "") for f in findings)),
        },
        # asyncio-APPROVED-1: to_thread wraps blocking pure-Python analysis function
        "top_files": await asyncio.to_thread(_file_influence, analyst_report, doc_index, top_n),
        # asyncio-APPROVED-1: to_thread wraps blocking pure-Python analysis function
        "source_type_x_finding_type": await asyncio.to_thread(_source_type_matrix, analyst_report),
        # asyncio-APPROVED-1: to_thread wraps blocking pure-Python analysis function
        "doc_type_x_finding_type": await asyncio.to_thread(_doc_type_matrix, analyst_report, doc_index),
        # asyncio-APPROVED-1: to_thread wraps blocking pure-Python analysis function
        "coverage": await asyncio.to_thread(_coverage_analysis, analyst_report, doc_list),
        # asyncio-APPROVED-1: to_thread wraps blocking pure-Python analysis function
        "investment_doc_coverage": await asyncio.to_thread(_per_investment_doc_coverage, doc_list),
        # asyncio-APPROVED-1: to_thread wraps blocking pure-Python analysis function
        "weak_findings": await asyncio.to_thread(_weak_findings, analyst_report),
        # asyncio-APPROVED-1: to_thread wraps blocking pure-Python analysis function
        "unresolved_gaps": await asyncio.to_thread(_unresolved_gaps, analyst_report),
        "cross_cutting": cross_cutting,
        "expected_docs": [],
        "nqpr_usage_diff": {},
    }

    # Write evidence_audit.json to run_dir/audit/
    if run_dir:
        out_dir = Path(run_dir) / "audit"

        def _write_audit_json():
            out_dir.mkdir(parents=True, exist_ok=True)
            (out_dir / "evidence_audit.json").write_text(
                json.dumps(_strip_internal_keys(audit), indent=2, default=str),
                encoding="utf-8",
            )

        # asyncio-APPROVED-1: to_thread wraps blocking JSON file write
        await asyncio.to_thread(_write_audit_json)

    return {"audit": audit}


async def write_brief(state: EvidenceAuditState) -> dict:
    audit = state.get("audit") or {}
    if not audit:
        return {"errors": ["run_audit must complete before write_brief"]}

    # asyncio-APPROVED-1: to_thread wraps blocking pure-Python brief formatting function
    brief_md = await asyncio.to_thread(_format_brief, audit)
    run_dir = state.get("run_dir") or ""
    brief_path: Optional[str] = None

    if run_dir:
        out_dir = Path(run_dir) / "audit"

        def _write():
            out_dir.mkdir(parents=True, exist_ok=True)
            p = out_dir / "team_brief.md"
            p.write_text(brief_md, encoding="utf-8")
            return str(p)

        # asyncio-APPROVED-1: to_thread wraps blocking file write
        brief_path = await asyncio.to_thread(_write)

    return {"brief_md": brief_md, "brief_path": brief_path}


async def write_workbook(state: EvidenceAuditState) -> dict:
    """Write xlsx workbook. No-op when output_xlsx is False."""
    if not state.get("output_xlsx"):
        return {}

    audit = state.get("audit") or {}
    run_dir = state.get("run_dir") or ""
    if not audit or not run_dir:
        return {}

    def _write() -> str:
        import openpyxl
        wb = openpyxl.Workbook()

        # Summary sheet
        ws = wb.active
        ws.title = "Summary"
        summary = audit.get("summary", {})
        for row_idx, (k, v) in enumerate(summary.items(), start=1):
            ws.cell(row=row_idx, column=1, value=k)
            ws.cell(row=row_idx, column=2, value=str(v))

        # Top files sheet
        ws2 = wb.create_sheet("Top Files")
        top_files = audit.get("top_files", [])
        if top_files:
            headers = list(top_files[0].keys())
            for col, h in enumerate(headers, start=1):
                ws2.cell(row=1, column=col, value=h)
            for row_idx, row in enumerate(top_files, start=2):
                for col, h in enumerate(headers, start=1):
                    val = row.get(h)
                    if isinstance(val, (dict, list)):
                        val = json.dumps(val, default=str)[:200]
                    ws2.cell(row=row_idx, column=col, value=val)

        # Coverage sheet
        ws3 = wb.create_sheet("Never Cited")
        never_cited = audit.get("coverage", {}).get("never_cited", [])
        if never_cited:
            headers3 = list(never_cited[0].keys())
            for col, h in enumerate(headers3, start=1):
                ws3.cell(row=1, column=col, value=h)
            for row_idx, row in enumerate(never_cited, start=2):
                for col, h in enumerate(headers3, start=1):
                    ws3.cell(row=row_idx, column=col, value=str(row.get(h, "")))

        out_dir = Path(run_dir) / "audit"
        out_dir.mkdir(parents=True, exist_ok=True)
        path = out_dir / "evidence_audit.xlsx"
        wb.save(str(path))
        return str(path)

    try:
        # asyncio-APPROVED-1: to_thread wraps blocking openpyxl workbook write
        workbook_path = await asyncio.to_thread(_write)
        return {"workbook_path": workbook_path}
    except Exception as exc:
        return {"errors": [f"write_workbook failed: {exc}"]}


async def write_diagnosis(state: EvidenceAuditState) -> dict:
    """Write detailed diagnosis JSON. No-op when output_diagnosis is False."""
    if not state.get("output_diagnosis"):
        return {}

    audit = state.get("audit") or {}
    run_dir = state.get("run_dir") or ""
    if not audit or not run_dir:
        return {}

    # Diagnosis = detailed weak-finding + coverage gap report
    diagnosis = {
        "program": audit.get("program"),
        "weak_findings": audit.get("weak_findings", []),
        "never_cited": audit.get("coverage", {}).get("never_cited", []),
        "investment_doc_coverage": audit.get("investment_doc_coverage", {}),
        "unresolved_gaps": audit.get("unresolved_gaps", []),
    }

    def _write() -> str:
        out_dir = Path(run_dir) / "audit"
        out_dir.mkdir(parents=True, exist_ok=True)
        p = out_dir / "evidence_diagnosis.json"
        p.write_text(json.dumps(diagnosis, indent=2, default=str), encoding="utf-8")
        return str(p)

    # asyncio-APPROVED-1: to_thread wraps blocking JSON file write
    await asyncio.to_thread(_write)
    return {}


async def rollup(state: EvidenceAuditState) -> dict:
    """Cross-program rollup. No-op unless needed (for future multi-program support)."""
    return {}


# ---------------------------------------------------------------------------
# Graph compilation
# ---------------------------------------------------------------------------

_builder = StateGraph(EvidenceAuditState)
_builder.add_node("load_artifacts", load_artifacts)
_builder.add_node("run_audit", run_audit)
_builder.add_node("write_brief", write_brief)
_builder.add_node("write_workbook", write_workbook)
_builder.add_node("write_diagnosis", write_diagnosis)
_builder.add_node("rollup", rollup)

_builder.add_edge(START, "load_artifacts")
_builder.add_edge("load_artifacts", "run_audit")
_builder.add_edge("run_audit", "write_brief")
_builder.add_edge("write_brief", "write_workbook")
_builder.add_edge("write_workbook", "write_diagnosis")
_builder.add_edge("write_diagnosis", "rollup")
_builder.add_edge("rollup", END)

evidence_audit_graph = _builder.compile()


def create_audit_state(
    program: str,
    data_root: str,
    run_dir_name: str = "phaseA-data-prep-full",
    top_n_files: int = 25,
    skip_llm_expected_docs: bool = True,
    output_xlsx: bool = False,
    output_diagnosis: bool = False,
) -> dict:
    """Build a valid initial EvidenceAuditState dict."""
    return {
        "program": program,
        "data_root": data_root,
        "run_dir_name": run_dir_name,
        "top_n_files": top_n_files,
        "skip_llm_expected_docs": skip_llm_expected_docs,
        "output_xlsx": output_xlsx,
        "output_diagnosis": output_diagnosis,
        "analyst_report": None,
        "doc_list": None,
        "investment_scoring": None,
        "run_dir": None,
        "audit": None,
        "brief_md": None,
        "brief_path": None,
        "workbook_path": None,
        "rollup_md_path": None,
        "errors": [],
    }
